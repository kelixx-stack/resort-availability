"""
롯데리조트 잔여객실 크롤러
============================
- Playwright 기반
- SSO 로그인 처리
- 리조트별 날짜별 순차 조회 (3개월)
- bizCd: 속초=81, 부여=61, 김해=91, 제주=71

실행: python lotte_crawler.py
"""

import os, time, json, re
from datetime import datetime, date, timedelta, timezone
from dateutil.relativedelta import relativedelta
from dotenv import load_dotenv
from playwright.sync_api import sync_playwright
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ─── 설정 ─────────────────────────────────────────────────────
load_dotenv(dotenv_path=os.path.join(os.path.dirname(__file__), '..', '.env'))

LOTTE_ID = os.getenv("LOTTE_ID", "jbhahaha1@naver.com")
LOTTE_PW = os.getenv("LOTTE_PW", "sjdhkdml2@")

LOGIN_URL   = "https://www.lottehotel.com/global/ko/login/rewards"
RESERVE_URL = "https://resort.lottehotel.com/main/ko/reservation/accommodation"

# 수집 대상 리조트 (bizCd → 리조트명)
TARGET_RESORTS = {
    "81": "롯데리조트 속초",
    "61": "롯데리조트 부여",
    "91": "롯데호텔앤리조트 김해",
}

START_DATE   = date.today()
MONTHS_COUNT = 3
KEEP_DAYS    = 7

OUTPUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data")
os.makedirs(OUTPUT_DIR, exist_ok=True)

TIMESTAMP  = datetime.now(timezone(timedelta(hours=9))).strftime("%Y%m%d_%H%M%S")
EXCEL_FILE = os.path.join(OUTPUT_DIR, f"lotte_{TIMESTAMP}.xlsx")
TXT_FILE   = os.path.join(OUTPUT_DIR, f"lotte_{TIMESTAMP}.txt")

WEEKDAYS = ["월", "화", "수", "목", "금", "토", "일"]

# ─── 날짜 목록 ────────────────────────────────────────────────
def build_date_range():
    end = START_DATE + relativedelta(months=MONTHS_COUNT)
    dates, cur = [], START_DATE
    while cur < end:
        dates.append(cur)
        cur += timedelta(days=1)
    return dates

# ─── 로그인 ───────────────────────────────────────────────────
def login(page):
    print("  로그인 페이지 접속...")
    for attempt in range(3):
        try:
            page.goto(LOGIN_URL, timeout=60000, wait_until="domcontentloaded")
            break
        except Exception as e:
            if attempt == 2:
                raise e
            print(f"  [경고] 로그인 페이지 로드 시도 {attempt+1} 실패. 재시도 중... ({e})")
            page.wait_for_timeout(3000)

    # 페이지 렌더링 대기
    page.wait_for_timeout(6000)

    # 1. 개인정보 이용동의 체크박스 클릭
    try:
        print("  개인정보 이용동의 체크박스 선택...")
        page.evaluate("document.querySelector('#personal-info-agree-check').click()")
        page.wait_for_timeout(1000)
    except Exception as e:
        print(f"  [경고] 체크박스 선택 실패: {e}")

    # 2. 개인정보 동의 '동의하고 통합 서비스 이용하기' 버튼 클릭
    try:
        print("  개인정보 이용동의 버튼 클릭...")
        page.evaluate("""() => {
            const btn = Array.from(document.querySelectorAll('button')).find(b => b.textContent.includes('동의하고') || b.textContent.includes('Agree'));
            if (btn) btn.click();
        }""")
        page.wait_for_timeout(2000)
    except Exception as e:
        print(f"  [경고] 이용동의 버튼 클릭 실패: {e}")

    # 3. 쿠키 동의 '전체 동의' 버튼 클릭
    try:
        print("  쿠키 동의 '전체 동의' 버튼 클릭...")
        page.evaluate("""() => {
            const btn = Array.from(document.querySelectorAll('button')).find(b => b.textContent.includes('전체 동의') || b.textContent.includes('Accept All'));
            if (btn) btn.click();
        }""")
        page.wait_for_timeout(2000)
    except Exception as e:
        print(f"  [경고] 쿠키 동의 버튼 클릭 실패: {e}")

    # 4. L.POINT 로그인 탭 선택
    try:
        print("  L.POINT 로그인 탭 선택...")
        page.evaluate("""() => {
            const tab = Array.from(document.querySelectorAll('button, a, span')).find(el => el.textContent === 'L.POINT 로그인');
            if (tab) tab.click();
        }""")
        page.wait_for_timeout(2000)
    except Exception as e:
        print(f"  [경고] L.POINT 탭 선택 실패: {e}")

    # 5. 계정 정보 입력
    try:
        print("  계정 정보 입력...")
        page.fill("input[name='loginId'] >> visible=true", LOTTE_ID)
        page.fill("input[name='loginPw'] >> visible=true", LOTTE_PW)
    except Exception as e:
        print(f"  [경고] 계정 입력 실패: {e}")
        raise e

    # 6. 로그인 버튼 클릭
    try:
        print("  로그인 버튼 클릭...")
        page.click("button.btn-cont-64", force=True)
        page.wait_for_timeout(8000)
    except Exception as e:
        print(f"  [경고] 로그인 버튼 클릭 실패: {e}")
        raise e

    print(f"  [성공] 로그인 완료 → {page.url}")

def clean_room_name(name):
    import re
    return re.sub(r"^\[.*?\]\s*", "", name).strip()

def collect_all(page):
    print("  [정보] 롯데 고속 API 잔여객실 수집기 구동...")
    
    # 롯데리조트 메인 페이지로 진입하여 리조트 subdomain 세션 및 쿠키 활성화
    print("  [정보] 리조트 서브도메인 세션 활성화 중...")
    for attempt in range(3):
        try:
            page.goto("https://resort.lottehotel.com/main/ko/index", timeout=60000, wait_until="domcontentloaded")
            break
        except Exception as e:
            if attempt == 2:
                raise e
            print(f"  [경고] 리조트 메인 페이지 로드 시도 {attempt+1} 실패. 재시도 중... ({e})")
            page.wait_for_timeout(3000)
    page.wait_for_timeout(5000)
    
    # 수집 대상 날짜 범위 설정 (오늘부터 90일)
    dates = build_date_range()
    print(f"  [정보] 수집 대상 날짜 범위: {dates[0]} ~ {dates[-1]} (총 {len(dates)}일)")
        
    memberships = ["6124224400", "6124340800"]
    
    tasks = []
    for biz_cd, resort_nm in TARGET_RESORTS.items():
        for d in dates:
            checkin_str = d.strftime("%Y%m%d")
            checkout_str = (d + timedelta(days=1)).strftime("%Y%m%d")
            month_label = d.strftime("%Y.%m")
            day_str = str(d.day)
            
            for mem_no in memberships:
                url = (
                    f"https://resort.lottehotel.com/api/main/ko/reservation/roomList"
                    f"?rsvType=BAR&procType=&bizCd={biz_cd}&checkinDt={checkin_str}&checkoutDt={checkout_str}"
                    f"&memberNo={mem_no}&exclusiveCd=&rsvNo=&userId=&userNm=&userMobile="
                    f"&roomType=&roomFlg=&membYearUseDaysType=1&deadLineDay=&packageNo=&petDetailList="
                    f"&waitingPassYn=&refreshCouponNo=&ownType=5&roomCnt=1"
                )
                tasks.append({
                    "url": url,
                    "resort_nm": resort_nm,
                    "biz_cd": biz_cd,
                    "month_label": month_label,
                    "day_str": day_str,
                    "checkin_str": checkin_str,
                    "member_no": mem_no
                })
                
    print(f"  [정보] API 요청 대상 생성 완료 (총 {len(tasks)}개)")
    
    js_code = """
    async function fetchAllCalendar(tasks) {
        const results = [];
        const chunkSize = 24; // 24개 동시 요청으로 병렬 수집 성능 대폭 개선 (WAF 차단 회피 한도 유지)
        for (let i = 0; i < tasks.length; i += chunkSize) {
            const chunk = tasks.slice(i, i + chunkSize);
            const promises = chunk.map(task => 
                fetch(task.url)
                    .then(async r => {
                        if (r.status !== 200) {
                            return { task, success: false, error: 'Status ' + r.status };
                        }
                        const data = await r.json();
                        return { task, data, success: true };
                    })
                    .catch(err => ({ task, error: err.message, success: false }))
            );
            const chunkResults = await Promise.all(promises);
            results.push(...chunkResults);
        }
        return { results };
    }
    """
    
    t0 = time.time()
    res_obj = page.evaluate(
        f"async (args) => {{ {js_code}; return await fetchAllCalendar(args.tasks); }}",
        {"tasks": tasks}
    )
    print(f"  [정보] API 잔여객실 수집 완료 (소요시간: {time.time() - t0:.2f}초)")
    
    results = res_obj["results"]
    
    all_data = []
    collect_dt = datetime.now(timezone(timedelta(hours=9))).strftime("%Y-%m-%d %H:%M")
    
    for r in results:
        task = r["task"]
        if not r["success"]:
            # API 수집 실패 건은 무시하고 진행
            continue
            
        data = r["data"]
        rooms = data.get("roomList", [])
        
        for rm in rooms:
            avail_type = rm.get("availableRsvType", "")
            
            # 필터: 예약 가능 ('Y')
            if avail_type == "Y":
                avail_cnt = int(rm.get("roomCnt", 0))
                
                # 남은 객실 수 > 0
                if avail_cnt > 0:
                    dt_obj = datetime.strptime(task["checkin_str"], "%Y%m%d").date()
                    yoil = WEEKDAYS[dt_obj.weekday()]
                    
                    room_nm = clean_room_name(rm.get("roomNm", "객실"))
                        
                    price = rm.get("minRateAmt", "") or rm.get("roomAvgAmt", "")
                    
                    all_data.append({
                        "수집일시": collect_dt,
                        "브랜드": "롯데",
                        "리조트명": task["resort_nm"],
                        "지역": "",
                        "년월": task["month_label"],
                        "일": task["day_str"],
                        "요일": yoil,
                        "객실타입": room_nm,
                        "예약가능수": str(avail_cnt),
                        "요금": str(price)
                    })
                
    # 중복 제거 (지점별/날짜별/객실타입별)
    seen = set()
    deduped = []
    for row in all_data:
        key = (row["리조트명"], row["년월"], row["일"], row["객실타입"])
        if key not in seen:
            seen.add(key)
            deduped.append(row)
            
    print(f"  [정보] 중복 제거 전: {len(all_data)}건 -> 중복 제거 후: {len(deduped)}건")
    return deduped

# ─── Excel/TXT 저장 ───────────────────────────────────────────
COLS = ["수집일시", "브랜드", "리조트명", "지역", "년월", "일", "요일", "객실타입", "예약가능수", "요금"]
HEADER_FILL = PatternFill("solid", start_color="DC2626")
AVAIL_FILL  = PatternFill("solid", start_color="FEE2E2")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

def save_excel(all_data):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "롯데_예약가능"

    for c, h in enumerate(COLS, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = Font(bold=True, color="FFFFFF", name="맑은 고딕", size=10)
        cell.alignment = CENTER

    for r, row in enumerate(all_data, 2):
        for c, col in enumerate(COLS, 1):
            cell = ws.cell(row=r, column=c, value=row.get(col, ""))
            cell.fill = AVAIL_FILL
            cell.font = Font(name="맑은 고딕", size=9)
            cell.alignment = CENTER

    widths = [18, 8, 20, 10, 10, 6, 6, 30, 12, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(EXCEL_FILE)
    print(f"  [성공] Excel: {os.path.basename(EXCEL_FILE)}")

def save_txt(all_data):
    lines = [
        "="*60,
        "  롯데리조트 잔여객실 현황",
        f"  수집일시: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        f"  수집건수: {len(all_data)}건",
        "="*60,
    ]
    resort_map = {}
    for row in all_data:
        resort_map.setdefault(row["리조트명"], []).append(row)

    for resort, rows in sorted(resort_map.items()):
        lines.append(f"\n■ {resort}")
        for row in sorted(rows, key=lambda x: (x["년월"], x["일"].zfill(2))):
            lines.append(
                f"  {row['년월']}.{row['일']:>2}({row['요일']}) "
                f"{row['객실타입'] or '객실'} | {row['예약가능수']}실"
            )

    with open(TXT_FILE, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))
    print(f"  [성공] TXT : {os.path.basename(TXT_FILE)}")

def cleanup_old_files():
    import glob as gb
    import re
    now = datetime.now()
    for pattern in [os.path.join(OUTPUT_DIR, "lotte_*.xlsx"),
                    os.path.join(OUTPUT_DIR, "lotte_*.txt")]:
        for f in gb.glob(pattern):
            filename = os.path.basename(f)
            match = re.search(r"\d{8}", filename)
            if match:
                try:
                    file_date = datetime.strptime(match.group(0), "%Y%m%d")
                    if (now - file_date).days >= KEEP_DAYS:
                        os.remove(f)
                        print(f"  [삭제] 삭제: {filename}")
                except Exception:
                    pass

# ─── 메인 ─────────────────────────────────────────────────────
def main():
    print("\n" + "="*55)
    print("  롯데리조트 잔여객실 크롤러")
    print("="*55)
    print(f"  대상: {', '.join(TARGET_RESORTS.values())}")
    print(f"  기간: {MONTHS_COUNT}개월")

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        page = browser.new_page(viewport={"width": 1280, "height": 900})

        try:
            print("\n[1/4] 로그인...")
            login(page)

            print("\n[2/4] 잔여객실 수집...")
            all_data = collect_all(page)

        except Exception as e:
            import traceback
            print(f"\n[오류] 오류: {e}")
            traceback.print_exc()
            all_data = []
        finally:
            pass
            browser.close()

    if all_data:
        print("\n[3/4] 파일 저장...")
        save_excel(all_data)

        print("\n[4/4] 오래된 파일 정리...")
        cleanup_old_files()

        print(f"\n{'='*55}")
        print(f"  [성공] 완료! 총 {len(all_data)}건")
        print(f"  [결과] {os.path.basename(EXCEL_FILE)}")
        print("="*55)
    else:
        print("\n[경고] 수집 데이터 없음 — 로그인/페이지 구조 확인 필요")

if __name__ == "__main__":
    main()
