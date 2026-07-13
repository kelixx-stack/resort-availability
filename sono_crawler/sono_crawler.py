"""
소노호텔앤리조트(구 대명리조트) 잔여객실 초고속 API 크롤러
================================================
- Playwright 기반
- 로그인 세션 기반 대표회원번호 자동 조회
- 3개월간 일자별 병렬 API 쿼리 (단 10초 내외 소요)
- 수집 데이터 표준 규격 저장 (sono_YYYYMMDD_HHMM.xlsx)
"""

import os
import time
import json
import re
from datetime import datetime, date, timedelta, timezone
from dateutil.relativedelta import relativedelta
from dotenv import load_dotenv
from playwright.sync_api import sync_playwright
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ─── 설정 ─────────────────────────────────────────────────────
load_dotenv(dotenv_path=os.path.join(os.path.dirname(__file__), '..', '.env'))

SONO_ID = os.getenv("DAEMYUNG_ID", "sbssys")
SONO_PW = os.getenv("DAEMYUNG_PW", "hanains_2016")

LOGIN_URL   = "https://www.sonohotelsresorts.com/member/login"
RESERVE_URL = "https://www.sonohotelsresorts.com/reserve/remaining/pc"

START_DATE   = date.today()
MONTHS_COUNT = 3
KEEP_DAYS    = 7

OUTPUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data")
os.makedirs(OUTPUT_DIR, exist_ok=True)

TIMESTAMP  = datetime.now(timezone(timedelta(hours=9))).strftime("%Y%m%d_%H%M%S")
EXCEL_FILE = os.path.join(OUTPUT_DIR, f"sono_{TIMESTAMP}.xlsx")
TXT_FILE   = os.path.join(OUTPUT_DIR, f"sono_{TIMESTAMP}.txt")

WEEKDAYS = ["월", "화", "수", "목", "금", "토", "일"]

# ─── 날짜 목록 생성 ───────────────────────────────────────────
def build_date_range():
    end = START_DATE + relativedelta(months=MONTHS_COUNT)
    dates = []
    cur = START_DATE
    while cur < end:
        dates.append(cur)
        cur += timedelta(days=1)
    return dates

# ─── Excel 저장 ───────────────────────────────────────────────
COLS = ["수집일시", "브랜드", "리조트명", "지역", "년월", "일", "요일", "객실타입", "예약가능수"]
HEADER_FILL = PatternFill("solid", start_color="1A56DB")
AVAIL_FILL  = PatternFill("solid", start_color="EBF5FF")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

def save_excel(all_data):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "소노_예약가능"

    for c, h in enumerate(COLS, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = Font(bold=True, color="FFFFFF", name="맑은 고딕", size=10)
        cell.alignment = CENTER

    for r, row in enumerate(all_data, 2):
        for c, col in enumerate(COLS, 1):
            cell = ws.cell(row=r, column=c, value=row.get(col, ""))
            cell.fill = AVAIL_FILL
            cell.font = Font(name="맑은 고딕", size=9)
            cell.alignment = CENTER

    widths = [18, 8, 20, 10, 10, 6, 6, 30, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(EXCEL_FILE)
    print(f"  [성공] Excel: {os.path.basename(EXCEL_FILE)}")

# ─── TXT 저장 ─────────────────────────────────────────────────
def save_txt(all_data):
    lines = [
        "="*60,
        "  소노호텔앤리조트 잔여객실 현황",
        f"  수집일시: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        f"  수집건수: {len(all_data)}건",
        "="*60,
    ]
    resort_map = {}
    for row in all_data:
        resort_map.setdefault(row["리조트명"], []).append(row)

    for resort, rows in sorted(resort_map.items()):
        lines.append(f"\n■ {resort}")
        for row in sorted(rows, key=lambda x: (x["년월"], x["일"].zfill(2))):
            lines.append(
                f"  {row['년월']}.{row['일']:>2}({row['요일']}) "
                f"{row['객실타입'] or '객실'} | 예약가능수: {row['예약가능수']}실"
            )

    with open(TXT_FILE, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))
    print(f"  [성공] TXT : {os.path.basename(TXT_FILE)}")

# ─── 오래된 파일 정리 ─────────────────────────────────────────
def cleanup_old_files():
    import glob as gb
    import re
    now = datetime.now()
    for pattern in [os.path.join(OUTPUT_DIR, "sono_*.xlsx"),
                    os.path.join(OUTPUT_DIR, "sono_*.txt")]:
        for f in gb.glob(pattern):
            filename = os.path.basename(f)
            match = re.search(r"\d{8}", filename)
            if match:
                try:
                    file_date = datetime.strptime(match.group(0), "%Y%m%d")
                    if (now - file_date).days >= KEEP_DAYS:
                        os.remove(f)
                        print(f"  [삭제] 삭제: {filename}")
                except Exception:
                    pass

# ─── 메인 ─────────────────────────────────────────────────────
def main():
    print("\n" + "="*55)
    print("  소노호텔앤리조트 잔여객실 초고속 API 크롤러")
    print("="*55)

    dates = build_date_range()
    date_strs = [d.strftime("%Y%m%d") for d in dates]
    all_data = []

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        context = browser.new_context()
        page = context.new_page()

        try:
            print("\n[1/3] 로그인...")
            for attempt in range(3):
                try:
                    page.goto(LOGIN_URL, timeout=60000, wait_until="domcontentloaded")
                    break
                except Exception as e:
                    if attempt == 2:
                        raise e
                    print(f"  [경고] 로그인 페이지 로드 시도 {attempt+1} 실패. 재시도 중... ({e})")
                    page.wait_for_timeout(3000)
            
            # 비밀번호 입력 필드가 나타날 때까지 대기
            try:
                page.wait_for_selector("input[type='password']", timeout=20000)
            except Exception:
                pass
            page.wait_for_timeout(2000)
            
            # 아이디 입력
            id_filled = False
            for sel in ["input[placeholder*='아이디']", "input[placeholder*='ID']", "input[id='id']", "input[name='id']"]:
                try:
                    el = page.query_selector(sel)
                    if el and el.is_visible():
                        el.fill(SONO_ID)
                        id_filled = True
                        break
                except Exception:
                    pass
            if not id_filled:
                page.locator("input[type='text']").first.fill(SONO_ID)

            # 비밀번호 입력
            page.locator("input[type='password']").first.fill(SONO_PW)
            page.keyboard.press("Enter")
            page.wait_for_load_state("networkidle")
            page.wait_for_timeout(3000)
            print(f"  [성공] 로그인 완료 → {page.url}")

            print("\n[2/3] 잔여객실 수집...")
            # 예약 페이지 진입
            for attempt in range(3):
                try:
                    page.goto(RESERVE_URL, timeout=60000, wait_until="domcontentloaded")
                    break
                except Exception as e:
                    if attempt == 2:
                        raise e
                    print(f"  [경고] 예약 페이지 로드 시도 {attempt+1} 실패. 재시도 중... ({e})")
                    page.wait_for_timeout(3000)
            page.wait_for_timeout(3000)

            # JS 코드를 브라우저 컨텍스트 내에서 실행하여 병렬 API 쿼리
            js_code = """
            async function fetchSono(dateStrings) {
                // 1. 대표 회원 번호 조회
                const memRes = await fetch("/api/hms/user/management/member/reserve/representativeMemberNo?lang=ko&deviceType=MO&mobileAppYn=N").then(r => r.json());
                const memNo = memRes.body;
                
                const results = [];
                const chunkSize = 15; // 15개 단위 병렬
                for (let i = 0; i < dateStrings.length; i += chunkSize) {
                    const chunk = dateStrings.slice(i, i + chunkSize);
                    const promises = chunk.map(cin_str => {
                        const year = parseInt(cin_str.substring(0, 4));
                        const month = parseInt(cin_str.substring(4, 6)) - 1;
                        const day = parseInt(cin_str.substring(6, 8));
                        const cinDate = new Date(year, month, day);
                        const coutDate = new Date(cinDate.getTime() + 24 * 60 * 60 * 1000);
                        const cout_str = coutDate.getFullYear() + 
                                         String(coutDate.getMonth() + 1).padStart(2, '0') + 
                                         String(coutDate.getDate()).padStart(2, '0');
                        
                        const url = `/api/hms/user/memberReservation/room/list/remaining?lang=ko&deviceType=MO&mobileAppYn=N` +
                                    `&memNo=${memNo}&userIndCd=Y&rsvIndCd=9` +
                                    `&ciYmd=${cin_str}&coYmd=${cout_str}&nights=1&rmCnt=1&adultCnt=1&childCnt=0` +
                                    `&rmTypeCode&paymentType&outsMemNo` +
                                    `&availableRsvDate=20270228&availableRsvDateForJeju=20270715&availableRsvDateForOuts=20270228&outsBrandSeq=-1`;
                                    
                        return fetch(url)
                            .then(r => r.json().then(data => ({ date: cin_str, data, success: true })))
                            .catch(err => ({ date: cin_str, error: err.message, success: false }));
                    });
                    const chunkResults = await Promise.all(promises);
                    results.push(...chunkResults);
                }
                return results;
            }
            """
            
            t0 = time.time()
            print(f"  총 {len(date_strs)}일의 날짜별 병렬 조회 요청 중...")
            api_results = page.evaluate(f"async (dates) => {{ {js_code}; return await fetchSono(dates); }}", date_strs)
            
            # 응답 데이터 파싱
            for r in api_results:
                if not r["success"]:
                    continue
                body = r["data"].get("body", [])
                if not body:
                    continue
                    
                date_str = r["date"]
                year_month = f"{date_str[:4]}.{date_str[4:6]}"
                day_str = str(int(date_str[6:8]))
                checkin_dt = date(int(date_str[:4]), int(date_str[4:6]), int(date_str[6:8]))
                
                for store in body:
                    store_nm = store.get("storeNm", "").strip()
                    for rt in store.get("rmTypeList", []):
                        status_cd = rt.get("rsvStatusCd")
                        # 예약가능("A" = 예약원활, "E" = 마감임박) 한 것만 수집
                        if status_cd in ["A", "E"]:
                            all_data.append({
                                "수집일시": datetime.now(timezone(timedelta(hours=9))).strftime("%Y-%m-%d %H:%M"),
                                "브랜드": "소노",
                                "리조트명": store_nm or "소노",
                                "지역": "",
                                "년월": year_month,
                                "일": day_str,
                                "요일": WEEKDAYS[checkin_dt.weekday()],
                                "객실타입": rt.get("roomTypeNm", "").strip(),
                                "예약가능수": str(rt.get("rsvRmCnt", 1)),
                            })
            # 중복 제거 (지점별/날짜별/객실타입별로 가장 큰 예약가능수를 가진 것 유지)
            dedup_map = {}
            for row in all_data:
                key = (row["리조트명"], row["년월"], row["일"], row["객실타입"])
                val = int(row["예약가능수"])
                if key not in dedup_map or val > int(dedup_map[key]["예약가능수"]):
                    dedup_map[key] = row
            all_data = list(dedup_map.values())
            
            print(f"  [성공] API 수집 완료 (소요시간: {time.time() - t0:.2f}초, 수집 건수: {len(all_data)}건)")

        except Exception as e:
            import traceback
            print(f"\n[오류] 오류: {e}")
            traceback.print_exc()
        finally:
            browser.close()

    if all_data:
        print("\n[3/3] 파일 저장...")
        save_excel(all_data)
        cleanup_old_files()

        print(f"\n{'='*55}")
        print(f"  [성공] 완료! 총 {len(all_data)}건")
        print(f"  [결과] {os.path.basename(EXCEL_FILE)}")
        print("="*55)
    else:
        print("\n[경고] 수집 데이터 없음 — API 구조 확인 필요")

if __name__ == "__main__":
    main()
