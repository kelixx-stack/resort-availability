"""
한화리조트 객실 예약 현황 수집기 v14
- Selenium Network 인터셉트 방식 (전체 자동화)
- 전체 리조트 탭 자동 순회 (BRCH_CD 불필요)
- data 폴더 자동 저장
- 30일 이상 파일 자동 삭제
- 표준 컬럼: 수집일시|브랜드|리조트명|년월|일|요일|객실타입명|예약가능수
"""

import json
import time
import os
import glob
from datetime import datetime, date, timezone, timedelta
from dateutil.relativedelta import relativedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service

# D:\휴양소\.env 경로에서 로드
load_dotenv(dotenv_path=os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', '.env'))

LOGIN_ID       = os.getenv("HANHWA_ID", "여기에_아이디")
LOGIN_PWD      = os.getenv("HANHWA_PW", "여기에_비밀번호")
MEMBERSHIP_PWD = os.getenv("HANHWA_MEMBERSHIP_PW", "여기에_회원권비밀번호")

# ★ 수집할 리조트 탭 이름 목록 (브라우저에 표시되는 이름 그대로)
# 전체 수집: 아래 목록 그대로 사용
# 일부만 수집: 원하는 항목만 남기면 됨
RESORT_TABS = [
    "설악 쏘라노",
    "설악 별관",
    "용인베잔송",
    "산정호수 안시",
    "대천 파로스",
    "경주 에톤",
    "경주 담톤",
    "제주",
    "평창",
    "거제 벨버디어",
    "거제 르 씨엘",
    "여수 벨메르",
    "더플라자 호텔",
    "브리드호텔 양양",
    "마티에 오시리아",
    "해운대",
]

START_DATE   = date.today().replace(day=1)  # 이번 달 1일부터
MONTHS_COUNT = 3                             # 수집 개월 수
KEEP_DAYS    = 7                            # 파일 보관 기간 (일)

# data 폴더 자동 생성
OUTPUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data")
os.makedirs(OUTPUT_DIR, exist_ok=True)

TIMESTAMP  = datetime.now(timezone(timedelta(hours=9))).strftime("%Y%m%d_%H%M%S")
EXCEL_FILE = os.path.join(OUTPUT_DIR, f"hanwha_{TIMESTAMP}.xlsx")
TXT_FILE   = os.path.join(OUTPUT_DIR, f"hanwha_{TIMESTAMP}.txt")

HNR_BASE     = "https://www.hanwharesort.co.kr"
BOOKING_BASE = "https://booking.hanwharesort.co.kr"
WEEKDAYS     = ["월", "화", "수", "목", "금", "토", "일"]

# ================================================================
# 1. Selenium 드라이버 생성
# ================================================================

def create_driver():
    options = Options()
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option("useAutomationExtension", False)
    options.add_experimental_option("prefs", {
        "credentials_enable_service": False,
        "profile.password_manager_enabled": False,
    })
    options.set_capability("goog:loggingPrefs", {"performance": "ALL"})
    options.add_argument("--headless=new")  # 백그라운드 실행 활성화

    driver = webdriver.Chrome(
        service=Service(ChromeDriverManager().install()),
        options=options
    )
    driver.execute_script(
        "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"
    )
    driver.set_script_timeout(30)
    driver.execute_cdp_cmd("Network.enable", {})
    return driver

# ================================================================
# 2. 2단계 로그인
# ================================================================

def do_login(driver):
    wait = WebDriverWait(driver, 20)

    print("  → 로그인 페이지 접속...")
    for attempt in range(3):
        try:
            driver.get(f"{HNR_BASE}/irsweb/resort3/member/login.do")
            break
        except Exception as e:
            if attempt == 2:
                raise e
            print(f"  ⚠ 로그인 페이지 접속 시도 {attempt+1} 실패. 재시도 중... ({e})")
            time.sleep(3)
    time.sleep(2)

    print("  → 1단계: 아이디/비밀번호 입력...")
    wait.until(EC.element_to_be_clickable((By.ID, "id"))).send_keys(LOGIN_ID)
    wait.until(EC.element_to_be_clickable((By.ID, "pwd"))).send_keys(LOGIN_PWD)
    wait.until(EC.element_to_be_clickable((By.ID, "btnLogin"))).click()
    time.sleep(3)
    print(f"     → {driver.current_url}")

    print("  → 2단계: 회원권 비밀번호 입력...")
    try:
        mem_pw = wait.until(EC.element_to_be_clickable(
            (By.CSS_SELECTOR,
             "input[name='membership_password'], input[id='membership_password']")
        ))
        mem_pw.clear()
        mem_pw.send_keys(MEMBERSHIP_PWD)
        time.sleep(0.5)
        wait.until(EC.element_to_be_clickable(
            (By.CSS_SELECTOR, "a.ui_button.brown")
        )).click()
        time.sleep(3)
        print(f"     → {driver.current_url}")
    except Exception as e:
        print(f"     ⚠ 2단계 오류: {e}")

    print("  → booking 사이트로 이동...")
    for attempt in range(3):
        try:
            driver.get(f"{BOOKING_BASE}/rst/rrs/0010/serviceM00.mvc?cont_no=")
            break
        except Exception as e:
            if attempt == 2:
                raise e
            print(f"  ⚠ booking 사이트 접속 시도 {attempt+1} 실패. 재시도 중... ({e})")
            time.sleep(3)
    time.sleep(5)
    print(f"     → {driver.current_url}")

    # booking에서도 2단계 인증 필요한 경우 처리
    try:
        mem_pw2 = WebDriverWait(driver, 5).until(EC.element_to_be_clickable(
            (By.CSS_SELECTOR,
             "input[name='membership_password'], input[id='membership_password']")
        ))
        print("  → booking 2단계 인증 감지! 비밀번호 재입력...")
        mem_pw2.clear()
        mem_pw2.send_keys(MEMBERSHIP_PWD)
        time.sleep(0.5)
        driver.find_element(By.CSS_SELECTOR, "a.ui_button.brown").click()
        time.sleep(4)
        print(f"     → {driver.current_url}")
    except Exception:
        pass

# ================================================================
# 3. 네트워크 로그에서 API 응답 추출
# ================================================================

def get_api_response_from_logs(driver):
    """검색 버튼 클릭 이후 로그만 남아있으므로 첫 번째 doExecute.mvc 응답 반환"""
    try:
        logs = driver.get_log("performance")
        for log in reversed(logs):
            msg = json.loads(log["message"])["message"]
            if msg.get("method") != "Network.responseReceived":
                continue
            url = msg.get("params", {}).get("response", {}).get("url", "")
            if "doExecute.mvc" not in url:
                continue
            request_id = msg["params"]["requestId"]
            try:
                body = driver.execute_cdp_cmd(
                    "Network.getResponseBody", {"requestId": request_id}
                )
                body_text = body.get("body", "")
                if body_text and body_text.strip().startswith("{"):
                    return body_text
            except Exception:
                continue
    except Exception as e:
        print(f"      ⚠ 로그 추출 오류: {e}")
    return None

# ================================================================
# 4. 날짜 헬퍼
# ================================================================

def fmt_date(date_str):
    try:
        d  = datetime.strptime(date_str, "%Y%m%d")
        wd = WEEKDAYS[d.weekday()]
        return d.strftime(f"%Y.%m.%d ({wd})")
    except Exception:
        return date_str

def fmt_month(d):
    return d.strftime("%Y년 %m월")

# ================================================================
# 5. 데이터 수집
# ================================================================

def collect_all_data(driver):
    all_data = []
    wait     = WebDriverWait(driver, 20)
    months   = [START_DATE + relativedelta(months=i) for i in range(MONTHS_COUNT)]
    total    = len(RESORT_TABS) * MONTHS_COUNT
    count    = 0

    for tab_name in RESORT_TABS:
        for ym in months:
            count += 1
            print(f"    [{count}/{total}] {tab_name} {fmt_month(ym)} 수집 중...")

            try:
                # ── 리조트 탭 클릭 (탭 이름으로 직접 찾기) ─────
                clicked = driver.execute_script(f"""
                    var tabs = document.querySelectorAll('ul#brch-tab li a, ul.calendar-tab li a');
                    for(var i=0; i<tabs.length; i++){{
                        if(tabs[i].textContent.trim() === '{tab_name}'){{
                            tabs[i].click(); return true;
                        }}
                    }}
                    return false;
                """)
                if not clicked:
                    print(f"           [경고] 탭 없음 (스킵): {tab_name}")
                    continue
                time.sleep(2)

                # ── 월 선택 드롭다운 (id="SCH_YM") ──────────────
                try:
                    sel_el = wait.until(EC.presence_of_element_located(
                        (By.ID, "SCH_YM")
                    ))
                    Select(sel_el).select_by_value(ym.strftime("%Y%m"))
                    time.sleep(0.5)
                except Exception as e:
                    print(f"           [오류] 월 선택 오류: {e}")

                # ── 검색 버튼 클릭 직전 로그 초기화 ─────────────
                driver.get_log("performance")

                # ── 검색 버튼 클릭 (id="calSch") ─────────────────
                try:
                    btn = wait.until(EC.element_to_be_clickable(
                        (By.ID, "calSch")
                    ))
                    driver.execute_script("arguments[0].click();", btn)
                except Exception as e:
                    print(f"           [오류] 검색 버튼 오류: {e}")

                time.sleep(3)

                # ── 네트워크 로그에서 API 응답 추출 ──────────────
                body = get_api_response_from_logs(driver)

                if body:
                    data      = json.loads(body)
                    records   = data.get("ds", {}).get("Data", {}).get("ds_result", [])
                    available = [r for r in records if r.get("RSRV_POSBL_YN") == "Y"]
                    print(f"           → {len(available)}건 예약가능")
                    for rec in available:
                        date_raw = rec.get("SESN_DATE", "")
                        all_data.append({
                            "날짜_raw":   date_raw,
                            "브랜드":     "한화",
                            "리조트명":   tab_name,
                            "년월":       fmt_yearmonth(date_raw),
                            "일":         fmt_day_only(date_raw),
                            "요일":       fmt_weekday(date_raw),
                            "객실타입명": rec.get("ROOM_TYPE_NM", ""),
                            "예약가능수": int(rec.get("RSRV_POSBL_CNT", 0)),
                        })
                else:
                    print(f"           [오류] API 응답 없음")

            except Exception as e:
                print(f"           [오류] 수집 오류: {e}")

            time.sleep(0.5)

    # 중복 제거
    seen = set()
    deduped = []
    for d in all_data:
        key = (d["브랜드"], d["리조트명"], d["날짜_raw"], d["객실타입명"])
        if key not in seen:
            seen.add(key)
            deduped.append(d)
    removed = len(all_data) - len(deduped)
    if removed:
        print(f"  [정보] 중복 {removed}건 제거 완료")
    return deduped

# ================================================================
# 6. Excel 저장
# ================================================================

AVAIL_FILL  = PatternFill("solid", start_color="C6EFCE")
HEADER_FILL = PatternFill("solid", start_color="2F4F8F")
SUB_FILL    = PatternFill("solid", start_color="4472C4")
CENTER      = Alignment(horizontal="center", vertical="center", wrap_text=True)

def hf(bold=True, size=10, color="FFFFFF"):
    return Font(bold=bold, color=color, name="맑은 고딕", size=size)

def df(size=9, color="000000"):
    return Font(name="맑은 고딕", size=size, color=color)

def fmt_yearmonth(date_str):
    """'20260501' → '2026.05'"""
    try:
        return f"{date_str[:4]}.{date_str[4:6]}"
    except Exception:
        return date_str

def fmt_day_only(date_str):
    """'20260501' → '01'"""
    try:
        return date_str[6:8]
    except Exception:
        return date_str

def fmt_weekday(date_str):
    """'20260501' → '금'"""
    try:
        d = datetime.strptime(date_str, "%Y%m%d")
        return WEEKDAYS[d.weekday()]
    except Exception:
        return ""

def save_excel(all_data):
    wb = openpyxl.Workbook()
    collect_dt = datetime.now(timezone(timedelta(hours=9))).strftime("%Y-%m-%d %H:%M")

    ws = wb.active
    ws.title = "전체현황(예약가능)"

    # 표준 컬럼: 수집일시 | 브랜드 | 리조트명 | 년월 | 일 | 요일 | 객실타입명 | 예약가능수
    cols = ["수집일시", "브랜드", "리조트명", "년월", "일", "요일", "객실타입명", "예약가능수"]
    for c, h in enumerate(cols, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = HEADER_FILL; cell.font = hf(); cell.alignment = CENTER
    ws.row_dimensions[1].height = 22

    for r, row in enumerate(all_data, 2):
        values = [
            collect_dt,
            row["브랜드"],
            row["리조트명"],
            row["년월"],
            row["일"],
            row["요일"],
            row["객실타입명"],
            row["예약가능수"],
        ]
        for c, v in enumerate(values, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.fill = AVAIL_FILL; cell.font = df(); cell.alignment = CENTER

    for i, w in enumerate([18, 10, 15, 10, 6, 6, 32, 12], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(EXCEL_FILE)
    print(f"  [성공] Excel: {os.path.basename(EXCEL_FILE)}")

# ================================================================
# 7. TXT 저장
# ================================================================

def save_txt(all_data):
    sep  = "=" * 72
    sep2 = "─" * 72
    lines = [
        sep, "  한화리조트 예약가능 객실 현황",
        f"  수집일시 : {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        f"  수집기간 : {fmt_month(START_DATE)} ~ "
        f"{fmt_month(START_DATE + relativedelta(months=MONTHS_COUNT-1))}",
        f"  대상리조트: {', '.join(RESORT_TABS)}", sep,
    ]
    for tab_name in RESORT_TABS:
        rdata = [d for d in all_data if d["리조트명"] == tab_name]
        lines += ["", sep2, f"  ■ {tab_name}", sep2]
        if not rdata:
            lines.append("  → 예약가능 객실 없음")
            continue
        for dt_raw in sorted(set(d["날짜_raw"] for d in rdata)):
            d0 = [x for x in rdata if x["날짜_raw"] == dt_raw][0]
            lines.append(f"\n  [{d0['년월']} {d0['일']} ({d0['요일']})]")
            for d in [x for x in rdata if x["날짜_raw"] == dt_raw]:
                lines.append(
                    f"    ○ {d['객실타입명']:<30} | "
                    f"예약가능: {d['예약가능수']:>4}실"
                )
    lines += ["", sep, f"  총 예약가능 건수: {len(all_data):,}건", sep, ""]
    with open(TXT_FILE, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))
    print(f"  [성공] TXT  : {os.path.basename(TXT_FILE)}")

# ================================================================
# 8. 오래된 파일 자동 삭제
# ================================================================

def cleanup_old_files():
    """KEEP_DAYS일 이상 된 hanwha_*.xlsx / *.txt 파일 삭제"""
    import re
    now = datetime.now()
    patterns = [
        os.path.join(OUTPUT_DIR, "hanwha_*.xlsx"),
        os.path.join(OUTPUT_DIR, "hanwha_*.txt"),
    ]
    deleted = 0
    for pattern in patterns:
        for fpath in glob.glob(pattern):
            filename = os.path.basename(fpath)
            match = re.search(r"\d{8}", filename)
            if match:
                try:
                    file_date = datetime.strptime(match.group(0), "%Y%m%d")
                    if (now - file_date).days >= KEEP_DAYS:
                        os.remove(fpath)
                        print(f"  [삭제] 삭제: {filename}")
                        deleted += 1
                except Exception as e:
                    print(f"  [오류] 삭제 실패: {filename} ({e})")
    if deleted:
        print(f"  [정보] 총 {deleted}개 파일 삭제 완료 ({KEEP_DAYS}일 이상)")
    else:
        print(f"  [정보] 삭제할 파일 없음 ({KEEP_DAYS}일 기준)")

# ================================================================
# 9. 메인
# ================================================================

if __name__ == "__main__":
    print("\n" + "=" * 55)
    print("  한화리조트 예약가능 객실 현황 수집기 v14")
    print("=" * 55)
    print(f"  리조트: {len(RESORT_TABS)}개 전체")
    print(f"  기간  : {MONTHS_COUNT}개월 ({fmt_month(START_DATE)} ~)")
    print(f"  저장  : {OUTPUT_DIR}")
    print("=" * 55 + "\n")

    t0 = time.time()
    driver = create_driver()

    try:
        print("[1/4] 로그인 중...")
        do_login(driver)

        print(f"\n[2/4] 객실 현황 수집 중...")
        all_data = collect_all_data(driver)
        print(f"  [정보] 예약가능 총 {len(all_data):,}건 수집 완료\n")

    finally:
        driver.quit()
        print("  [정보] 브라우저 종료\n")

    print("[3/4] 파일 저장 중...")
    save_excel(all_data)

    print("\n[4/4] 오래된 파일 정리 중...")
    cleanup_old_files()

    elapsed = time.time() - t0
    print(f"\n{'=' * 55}")
    print(f"  [성공] 완료!  소요시간: {elapsed:.1f}초")
    print(f"  [엑셀] {os.path.basename(EXCEL_FILE)}")
    print("=" * 55 + "\n")
